import openpyxl
from matplotlib import pyplot as plt
import numpy as np

#dados foram extraidos de: https://www.corona.ma.gov.br/
def pegaColuna(nome_file,i,j):
    arquivo = openpyxl.load_workbook('.\\'+nome_file) #carrega o arquivo
    coluna_geral=[]
    coluna_data=[]
    coluna_casos_novos=[]
    coluna_casos_acumulados=[]
    folha=arquivo.active
    maximo_linhas=folha.max_row

    for linha in range(i,maximo_linhas+1):
        celula=folha.cell(row=linha,column=j)
        if celula.value!= None:
            coluna_geral.append(celula.value.split(','))
    for elemento in coluna_geral:
        coluna_data.append(elemento[0])
        coluna_casos_novos.append(float(elemento[1]))
        coluna_casos_acumulados.append(float(elemento[2]))
            
    return coluna_data,coluna_casos_novos

def mediaMovel(dados,k):
    medias_moveis=[]
    for elemento in range (0,len(dados)):
        grupo=dados[elemento: elemento + k]
        if len(grupo) == k:
            
            media=sum(grupo)/k
            medias_moveis.append(media)
    return medias_moveis

def variacao(vetor):
    aux=vetor[(len(vetor)-15)::]
    soma=0
    for i in range(0,len(aux)-1):
        soma+=aux[i]
    media=soma/(len(aux)-1)
    v= ((aux[-1] - media)/media)*100
    return v


k=7
k1=14
dados=pegaColuna('DADOSCOVIDMARANHAO.xlsx',2,1)
medias=mediaMovel(dados[1],k)
medias1=mediaMovel(dados[1],k1)
data=dados[0]
v=variacao(medias)
print(v)


print('Saída: ',medias) #Resultado da saida

if v >=-15 and v <=15:
    print('ESTÁVEL')

if v >15:
    print('ALTA')

if v < -15:
    print('QUEDA')
    



plt.plot(np.arange(0,len(medias)),medias,'--',color='blue',label='Média móvel de casos últimos '+str(k)+' dias')
plt.plot(np.arange(0,len(dados[1])),dados[1],alpha=0.5,color='red',label='Casos Novos')

plt.xlabel('20/03/2020 - 12/11/2021 (Quantidade de Dias)') #legenda eixo x
plt.ylabel('Quantidade de Casos') #legenda eixo y


plt.grid()
plt.legend()
plt.title ('Casos de Covid-19 MA')
plt.show()


    
plt.plot(np.arange(0,len(medias1)),medias1,'--',color='green',label='Média móvel de últimos '+str(k1)+' dias')
plt.plot(np.arange(0,len(dados[1])),dados[1],alpha=0.5,color='red',label='Casos Novos')
plt.xlabel('20/03/2020 - 12/11/2021 (Quantidade de Dias)') #legenda eixo x
plt.ylabel('Quantidade de Casos') #legenda eixo y


plt.grid()
plt.legend()
plt.title ('Casos de Covid-19 MA')
plt.show()
